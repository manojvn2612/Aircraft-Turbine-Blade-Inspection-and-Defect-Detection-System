import base64
import body_model
import cv2
import hrnet_model
import json
import numpy as np
import openpyxl
import os
import requests as req_lib
import shutil
import subprocess
import sys
import threading
import time
import traceback
import uuid
import zipfile
from datetime import datetime
from flask import Flask, jsonify, request, send_file, Response
from flask_cors import CORS
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.utils import secure_filename

app = Flask(__name__)
CORS(app)


# =============================================================================
# CONFIGURATION AND CONSTANTS
# =============================================================================

# Directory Structure
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR  = "outputs"
RETRAIN_DIR = "retrain_data"

APPROVED_DIR      = os.path.join(RETRAIN_DIR, "approved", "images")
DEFECTIVE_DIR     = os.path.join(RETRAIN_DIR, "defective", "images")
APPROVED_LBL      = os.path.join(RETRAIN_DIR, "approved", "labels")
DEFECTIVE_LBL     = os.path.join(RETRAIN_DIR, "defective", "labels")
RETRAIN_QUEUE_DIR = os.path.join(RETRAIN_DIR, "retrain", "images")
RETRAIN_QUEUE_LBL = os.path.join(RETRAIN_DIR, "retrain", "labels")
UPLOAD_STAGING    = os.path.join(RETRAIN_DIR, "upload_staging", "images")
UPLOAD_ANNOT_LBL  = os.path.join(RETRAIN_DIR, "upload_staging", "labels")
LS_DATA_DIR       = os.path.join(RETRAIN_DIR, "label_studio_data")

SHOTS_FILE = os.path.join(BASE_DIR, "shots.json")

# Retrain Files
RETRAIN_LOG           = os.path.join(RETRAIN_DIR, "retrain_log.json")
RETRAIN_PROGRESS_FILE = os.path.join(RETRAIN_DIR, "retrain_progress.json")
RETRAIN_STOP_FILE     = os.path.join(RETRAIN_DIR, "retrain_stop.signal")
RETRAIN_SCRIPT        = os.path.join(BASE_DIR, "BFL_retrain.py")
RETRAIN_LOG_FILE      = os.path.join(RETRAIN_DIR, "retrain_stdout.log")

# File Extensions
ALLOWED_EXTENSIONS      = {".jpg", ".jpeg", ".png"}
ALLOWED_EXTENSIONS_WIDE = {".jpg", ".jpeg", ".png", ".bmp", ".tiff", ".tif", ".webp"}

# Processing Constants
MIN_BLADE_CROP_PX = 60
EDGE_MARGIN_FRAC  = 0.04

# Label Studio Configuration
LS_PORT            = 8080
LS_URL             = f"http://localhost:{LS_PORT}"
LS_API_TOKEN       = ""
LS_USERNAME        = "admin@blade.local"
LS_PASSWORD        = "bladeinspect123"
LS_PROJECT_ID_FILE = os.path.join(RETRAIN_DIR, ".ls_project_id")
PROJECT_TITLE      = "Retrain"

# Defect Classes and Colors
DEFECT_CLASSES = [
    "Cutter marks and fish marks",
    "Fingerprints and stains",
    "Ink marks",
    "Jig Marks",
    "Machining Marks",
    "Overcut",
    "Pocket",
    "Scratches and Black spots",
]
DEFECT_CLASS_COLORS = {
    "Cutter marks and fish marks": "#EF4444",
    "Scratches and Black spots":   "#F97316",
    "Fingerprints and stains":     "#EAB308",
    "Ink marks":                   "#22C55E",
    "Jig Marks":                   "#06B6D4",
    "Machining Marks":             "#3B82F6",
    "Overcut":                     "#8B5CF6",
    "Pocket":                      "#EC4899",
}
DEFECT_CLASS_INDEX = {name: i for i, name in enumerate(DEFECT_CLASSES)}

# Default Shots Configuration
DEFAULT_SHOTS = [
    {"sr": 1,  "part": "Top leading edge",          "section": "Aerofoil", "elevation": 38,  "camAngle": 85,  "tableAngle": 55,  "zoom": 1.5, "focus": 2085, "flash": 22},
    {"sr": 2,  "part": "Top leading part",           "section": "Aerofoil", "elevation": 127, "camAngle": 105, "tableAngle": 78,  "zoom": 1.5, "focus": 2085, "flash": 22},
    {"sr": 3,  "part": "Top mid part",               "section": "Aerofoil", "elevation": 104, "camAngle": 102, "tableAngle": 115, "zoom": 1.5, "focus": 2085, "flash": 22},
    {"sr": 4,  "part": "Top trailing part",          "section": "Aerofoil", "elevation": 48,  "camAngle": 88,  "tableAngle": 199, "zoom": 1.5, "focus": 2085, "flash": 22},
    {"sr": 5,  "part": "Top trailing edge",          "section": "Aerofoil", "elevation": 45,  "camAngle": 85,  "tableAngle": 245, "zoom": 1.5, "focus": 2020, "flash": 22},
    {"sr": 6,  "part": "Top leading opp part",       "section": "Aerofoil", "elevation": 41,  "camAngle": 91,  "tableAngle": 28,  "zoom": 1.5, "focus": 2085, "flash": 22},
    {"sr": 7,  "part": "Top mid opp part",           "section": "Aerofoil", "elevation": 14,  "camAngle": 82,  "tableAngle": 280, "zoom": 1.5, "focus": 2085, "flash": 22},
    {"sr": 8,  "part": "Top trailing opp part",      "section": "Aerofoil", "elevation": 102, "camAngle": 102, "tableAngle": 260, "zoom": 1.5, "focus": 2085, "flash": 22, "note": "Change focus"},
    {"sr": 9,  "part": "Middle leading edge",        "section": "Aerofoil", "elevation": 190, "camAngle": 91,  "tableAngle": 56,  "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 10, "part": "Middle leading part",        "section": "Aerofoil", "elevation": 185, "camAngle": 79,  "tableAngle": 262, "zoom": 1.5, "focus": 2085, "flash": 11},
    {"sr": 11, "part": "Middle mid part",            "section": "Aerofoil", "elevation": 190, "camAngle": 91,  "tableAngle": 83,  "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 12, "part": "Middle trailing part",       "section": "Aerofoil", "elevation": 157, "camAngle": 85,  "tableAngle": 86,  "zoom": 1.5, "focus": 2085, "flash": 22},
    {"sr": 13, "part": "Middle trailing edge",       "section": "Aerofoil", "elevation": 200, "camAngle": 90,  "tableAngle": 232, "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 14, "part": "Middle leading opp part",    "section": "Aerofoil", "elevation": 208, "camAngle": 92,  "tableAngle": 304, "zoom": 1.5, "focus": 2085, "flash": 8},
    {"sr": 15, "part": "Middle mid opp part",        "section": "Aerofoil", "elevation": 135, "camAngle": 74,  "tableAngle": 9,   "zoom": 1.5, "focus": 2085, "flash": 8, "note": "Add natural light"},
    {"sr": 16, "part": "Middle trailing opp part",   "section": "Aerofoil", "elevation": 189, "camAngle": 92,  "tableAngle": 254, "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 17, "part": "Bottom trailing edge",       "section": "Aerofoil", "elevation": 190, "camAngle": 95,  "tableAngle": 330, "zoom": 1.5, "focus": 2085, "flash": 2, "note": "Reverse blade"},
    {"sr": 18, "part": "Bottom leading part",        "section": "Aerofoil", "elevation": 42,  "camAngle": 81,  "tableAngle": 282, "zoom": 1.5, "focus": 2085, "flash": 8},
    {"sr": 19, "part": "Bottom mid part",            "section": "Aerofoil", "elevation": 83,  "camAngle": 92,  "tableAngle": 119, "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 20, "part": "Bottom trailing part",       "section": "Aerofoil", "elevation": 22,  "camAngle": 53,  "tableAngle": 292, "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 21, "part": "Bottom leading edge",        "section": "Aerofoil", "elevation": 85,  "camAngle": 96,  "tableAngle": 132, "zoom": 1.5, "focus": 2025, "flash": 14},
    {"sr": 22, "part": "Bottom leading opp part",    "section": "Aerofoil", "elevation": 49,  "camAngle": 88,  "tableAngle": 148, "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 23, "part": "Bottom mid opp part",        "section": "Aerofoil", "elevation": 143, "camAngle": 101, "tableAngle": 161, "zoom": 1.5, "focus": 2081, "flash": 14},
    {"sr": 24, "part": "Bottom trailing opp part",   "section": "Aerofoil", "elevation": 221, "camAngle": 110, "tableAngle": 298, "zoom": 1.5, "focus": 2085, "flash": 14, "note": "Please adjust cloth on top"},
    {"sr": 25, "part": "Sensitive Zone Opp leading", "section": "Base",     "elevation": 320, "camAngle": 133, "tableAngle": 159, "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 26, "part": "Sensitive Zone trailing",    "section": "Base",     "elevation": 37,  "camAngle": 47,  "tableAngle": 219, "zoom": 1.5, "focus": 2119, "flash": 14, "note": "Rotate blade upside down, add gloves"},
    {"sr": 27, "part": "Critical point leading",     "section": "Base",     "elevation": 155, "camAngle": 116, "tableAngle": 138, "zoom": 1.5, "focus": 2085, "flash": 14, "note": "Rotate blade upside down"},
    {"sr": 28, "part": "Weld face Trailing Edge",    "section": "Base",     "elevation": 0,   "camAngle": 73,  "tableAngle": 333, "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 29, "part": "Stub flanks middle",         "section": "Base",     "elevation": 20,  "camAngle": 82,  "tableAngle": 84,  "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 30, "part": "Trailing Edge stub opp",     "section": "Base",     "elevation": 13,  "camAngle": 79,  "tableAngle": 203, "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 31, "part": "Stub flanks mid opp pocket", "section": "Base",     "elevation": 55,  "camAngle": 83,  "tableAngle": 297, "zoom": 1.5, "focus": 2085, "flash": 14},
    {"sr": 32, "part": "Leading Edge Stub flanks",   "section": "Base",     "elevation": None, "camAngle": None, "tableAngle": None, "zoom": None, "focus": None, "flash": None,
     "note": "Cannot capture — needs camera height increase", "disabled": True},
]


# =============================================================================
# GLOBAL STATE
# =============================================================================

# Retrain process
_retrain_process: subprocess.Popen | None = None
_retrain_lock = threading.Lock()

# Label Studio process
_ls_process = None
_ls_lock = threading.Lock()
_cached_ls_token = None
_cached_token_lock = threading.Lock()

# Camera (MJPEG frame shared memory)
_camera_process = None
_camera_lock = threading.Lock()
_camera_frame: bytes | None = None
_camera_meta: dict = {}

# Pic clicker capture session
_pic_clicker_process = None
_pic_clicker_folder  = None
_pic_clicker_lock    = threading.Lock()


# =============================================================================
# STARTUP HELPERS
# =============================================================================

def _ensure_directories():
    """Create all required directories."""
    for d in [
        OUTPUT_DIR, APPROVED_DIR, DEFECTIVE_DIR, APPROVED_LBL, DEFECTIVE_LBL,
        RETRAIN_QUEUE_DIR, RETRAIN_QUEUE_LBL, UPLOAD_STAGING, UPLOAD_ANNOT_LBL,
        LS_DATA_DIR,
    ]:
        os.makedirs(d, exist_ok=True)


def _clear_stale_files():
    """Remove leftover training signal files from previous runs."""
    for path in [RETRAIN_PROGRESS_FILE, RETRAIN_STOP_FILE]:
        if os.path.exists(path):
            try:
                os.remove(path)
                print(f"[Startup] Cleared stale file: {path}")
            except Exception as e:
                print(f"[Startup] Could not remove {path}: {e}")


def _load_shots() -> list:
    """Load shots config from disk; create defaults if missing."""
    if not os.path.exists(SHOTS_FILE):
        with open(SHOTS_FILE, "w") as f:
            json.dump(DEFAULT_SHOTS, f, indent=2)
        print(f"[Shots] Created default shots.json with {len(DEFAULT_SHOTS)} entries.")
    with open(SHOTS_FILE) as f:
        return json.load(f)


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def _count_staging() -> int:
    """Count images waiting in the upload staging area."""
    return len([
        f for f in os.listdir(UPLOAD_STAGING)
        if os.path.splitext(f)[1].lower() in ALLOWED_EXTENSIONS
    ])


def _count_images_in_dir(directory: str) -> int:
    """Count allowed image files in a directory."""
    if not os.path.isdir(directory):
        return 0
    return len([
        f for f in os.listdir(directory)
        if os.path.splitext(f)[1].lower() in ALLOWED_EXTENSIONS
    ])


def _audit_trainable(img_dir: str, lbl_dir: str) -> dict:
    """
    Walk img_dir and check each image for a matching non-empty label file.
    Returns counts used by both the retrain gate and the UI stats,
    so the two never disagree.
    """
    if not os.path.isdir(img_dir):
        return {"total": 0, "with_labels": 0, "without_labels": 0}

    total = with_labels = without_labels = 0
    for fn in os.listdir(img_dir):
        if os.path.splitext(fn)[1].lower() not in ALLOWED_EXTENSIONS:
            continue
        total += 1
        lbl_path = os.path.join(lbl_dir, os.path.splitext(fn)[0] + ".txt")
        if os.path.exists(lbl_path):
            try:
                if os.path.getsize(lbl_path) > 0:
                    with_labels += 1
                else:
                    without_labels += 1
            except Exception:
                without_labels += 1
        else:
            without_labels += 1

    return {"total": total, "with_labels": with_labels, "without_labels": without_labels}


def _append_log(filename: str, label: str, source: str = "inspection"):
    """Append one entry to the retrain activity log."""
    log = []
    if os.path.exists(RETRAIN_LOG):
        try:
            with open(RETRAIN_LOG) as f:
                log = json.load(f)
        except Exception:
            log = []
    log.append({
        "filename": filename,
        "label":    label,
        "time":     datetime.now().isoformat(),
        "source":   source,
    })
    with open(RETRAIN_LOG, "w") as f:
        json.dump(log, f, indent=2)


# =============================================================================
# IMAGE PROCESSING
# =============================================================================

def _translate_detections(detections: list, x1: int, y1: int) -> list:
    """Offset detection bounding boxes by (x1, y1) to map crop coords → full image."""
    translated = []
    for det in detections:
        d = dict(det)
        bx, by, bw, bh = d["bbox"]
        d["bbox"] = [bx + x1, by + y1, bw, bh]
        translated.append(d)
    return translated


def _filter_edge_detections(detections: list, crop_w: int, crop_h: int) -> list:
    """Drop detections whose centre falls within the edge margin of the crop."""
    margin_x = EDGE_MARGIN_FRAC * crop_w
    margin_y = EDGE_MARGIN_FRAC * crop_h
    kept = []
    for det in detections:
        bx, by, bw, bh = det["bbox"]
        cx = bx + bw / 2
        cy = by + bh / 2
        if margin_x < cx < crop_w - margin_x and margin_y < cy < crop_h - margin_y:
            kept.append(det)
    return kept


def _process_image_pipeline(image, filename: str = "image") -> dict | None:
    """
    Run the full body-detection + defect-prediction pipeline on one image.

    Returns a result dict, or None if the image yielded no body crop at all.
    The returned dict always includes 'body_message' (None or a warning string
    for the UI).
    """
    raw = body_model.detect_bodies(image)

    # body_model may return a plain list or a dict with an optional warning.
    if isinstance(raw, dict):
        body_message    = raw.get("message")
        body_detections = raw.get("detections", [])
    else:
        body_message    = None
        body_detections = raw

    if not body_detections:
        return None  # Genuine empty result — nothing to work with.

    full_output_image = image.copy()
    all_detections    = []
    body_confidence   = 0.0

    for detection in body_detections:
        cropped         = detection["cropped_image"]
        x1, y1, x2, y2 = detection["box"]
        crop_h, crop_w  = cropped.shape[:2]

        if crop_w < MIN_BLADE_CROP_PX or crop_h < MIN_BLADE_CROP_PX:
            continue

        body_confidence = max(body_confidence, detection["confidence"])
        vis_crop, raw_detections = hrnet_model.predict(cropped)
        if vis_crop is None:
            continue

        filtered   = _filter_edge_detections(raw_detections, crop_w, crop_h)
        translated = _translate_detections(filtered, x1, y1)

        vis_resized = cv2.resize(vis_crop, (x2 - x1, y2 - y1))
        full_output_image[y1:y2, x1:x2] = vis_resized
        all_detections.extend(translated)

    return {
        "predicted_image": full_output_image,
        "detections":      all_detections,
        "body_confidence": body_confidence,
        "filename":        filename,
        "body_message":    body_message,
    }


def _encode_image_b64(image) -> str:
    """Encode a BGR numpy image to a base-64 PNG string."""
    _, buffer = cv2.imencode('.png', image)
    return base64.b64encode(buffer).decode('utf-8')


# =============================================================================
# EXCEL EXPORT
# =============================================================================

def _save_excel(results_per_image: list, output_path: str):
    """Write detection results to a formatted Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Defect Report"

    col_headers = [
        "Image", "Body Confidence", "Defect Class",
        "Defect Confidence", "BBox (x,y,w,h)", "Accepted", "Blade Label",
    ]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(col_headers, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal="center")
    for col_letter, width in zip("ABCDEFG", [30, 18, 16, 20, 22, 12, 16]):
        ws.column_dimensions[col_letter].width = width

    green_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    orange_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    violet_fill = PatternFill(start_color="E9D5FF", end_color="E9D5FF", fill_type="solid")

    row = 2
    for item in results_per_image:
        filename    = item["filename"]
        body_conf   = item["body_confidence"]
        detections  = item["detections"]
        blade_label = item.get("blade_label")

        label_fill = (
            green_fill  if blade_label == "approved"  else
            red_fill    if blade_label == "defective" else
            violet_fill if blade_label == "retrain"   else
            orange_fill
        )
        label_val = blade_label.upper() if blade_label else "UNLABELED"

        if not detections:
            ws.cell(row=row, column=1, value=filename)
            ws.cell(row=row, column=2, value=round(body_conf, 4))
            ws.cell(row=row, column=3, value="No defects detected")
            c       = ws.cell(row=row, column=6, value="ACCEPTED")
            c.fill  = green_fill
            lc      = ws.cell(row=row, column=7, value=label_val)
            lc.fill = label_fill
            lc.font = Font(bold=True)
            lc.alignment = Alignment(horizontal="center")
            row += 1
            continue

        for det in detections:
            ws.cell(row=row, column=1, value=filename)
            ws.cell(row=row, column=2, value=round(body_conf, 4))
            ws.cell(row=row, column=3, value=det["class_name"])
            ws.cell(row=row, column=4, value=det["confidence"])
            x, y, w, h = det["bbox"]
            ws.cell(row=row, column=5, value=f"({x},{y},{w},{h})")
            acc        = det["accepted"]
            ac         = ws.cell(row=row, column=6, value="ACCEPTED" if acc else "REJECTED")
            ac.fill    = green_fill if acc else red_fill
            ac.font    = Font(bold=True)
            ac.alignment = Alignment(horizontal="center")
            lc         = ws.cell(row=row, column=7, value=label_val)
            lc.fill    = label_fill
            lc.font    = Font(bold=True)
            lc.alignment = Alignment(horizontal="center")
            row += 1

    ws.cell(row=row + 1, column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    wb.save(output_path)


# =============================================================================
# YOLO LABEL HELPERS
# =============================================================================

def _detections_to_yolo_lines(detections: list, img_w: int, img_h: int) -> list:
    """Convert detection dicts to YOLO polygon label lines (normalised coords)."""
    lines = []
    for det in detections:
        cls_idx = DEFECT_CLASS_INDEX.get(det.get("class_name", ""))
        if cls_idx is None:
            continue
        bx, by, bw, bh = det["bbox"]
        x1, y1 = bx / img_w,         by / img_h
        x2, y2 = (bx + bw) / img_w,  by / img_h
        x3, y3 = (bx + bw) / img_w,  (by + bh) / img_h
        x4, y4 = bx / img_w,         (by + bh) / img_h
        lines.append(
            f"{cls_idx} {x1:.6f} {y1:.6f} {x2:.6f} {y2:.6f} "
            f"{x3:.6f} {y3:.6f} {x4:.6f} {y4:.6f}"
        )
    return lines


def _write_yolo_label(detections: list, image_path: str, label_path: str):
    """Write a YOLO polygon .txt label file next to an image."""
    if not detections:
        open(label_path, "w").close()
        return
    img = cv2.imread(image_path)
    if img is None:
        open(label_path, "w").close()
        return
    h, w = img.shape[:2]
    lines = _detections_to_yolo_lines(detections, w, h)
    with open(label_path, "w") as f:
        f.write("\n".join(lines))


# =============================================================================
# CAMERA HELPERS
# =============================================================================

def _launch_camera_app():
    """Spawn pic_clicker.py as a subprocess."""
    global _camera_process
    cam_path = os.path.join(BASE_DIR, "pic_clicker.py")
    if not os.path.exists(cam_path):
        print("[Camera] pic_clicker.py not found — skipping auto-launch.")
        return
    print("[Camera] Launching pic_clicker.py...")
    _camera_process = subprocess.Popen([sys.executable, cam_path], cwd=BASE_DIR)
    print(f"[Camera] pic_clicker.py launched (PID {_camera_process.pid})")


def camera_state_update(pData: bytes, width: int, height: int, fmt: str = "rgb"):
    """
    Called by the camera integration to push a new frame into shared state.
    Accepts raw RGB bytes or pre-encoded JPEG.
    """
    global _camera_frame, _camera_meta
    try:
        if fmt == "jpeg":
            jpeg = pData if isinstance(pData, bytes) else bytes(pData)
        else:
            arr  = np.frombuffer(pData, dtype=np.uint8).reshape((height, width, 3))
            bgr  = cv2.cvtColor(arr, cv2.COLOR_RGB2BGR)
            _, buf = cv2.imencode('.jpg', bgr, [cv2.IMWRITE_JPEG_QUALITY, 80])
            jpeg = buf.tobytes()
        with _camera_lock:
            _camera_frame = jpeg
            _camera_meta  = {"width": width, "height": height}
    except Exception as e:
        print(f"[camera_state_update] Error encoding frame: {e}")


def _generate_mjpeg():
    """Generator that yields MJPEG boundary frames for the /camera-stream endpoint."""
    while True:
        with _camera_lock:
            frame = _camera_frame
        if frame:
            yield (
                b"--frame\r\n"
                b"Content-Type: image/jpeg\r\n\r\n" + frame + b"\r\n"
            )
        time.sleep(1 / 30)


# =============================================================================
# RETRAIN PROCESS HELPERS
# =============================================================================

def _start_retrain_process(cmd: list, env: dict) -> subprocess.Popen:
    """
    Launch the training script, routing stdout+stderr to a log file.
    Using a file avoids pipe-buffer deadlocks that can occur with PIPE.
    """
    global _retrain_process
    try:
        open(RETRAIN_LOG_FILE, 'w').close()
    except Exception:
        pass
    log_fh = open(RETRAIN_LOG_FILE, 'w', buffering=1)  # line-buffered
    _retrain_process = subprocess.Popen(cmd, env=env, cwd=BASE_DIR,
                                        stdout=log_fh, stderr=log_fh)
    return _retrain_process


# =============================================================================
# LABEL STUDIO HELPERS
# =============================================================================

def _ls_is_running() -> bool:
    """Return True when Label Studio responds healthy on its health endpoint."""
    try:
        r = req_lib.get(f"{LS_URL}/health", timeout=3)
        return r.status_code == 200 and r.json().get("status") == "UP"
    except Exception:
        return False


def _ls_fetch_token_via_session() -> str:
    """Obtain an API token by logging in through the Label Studio web UI."""
    try:
        session    = req_lib.Session()
        login_page = session.get(f"{LS_URL}/user/login/", timeout=10)
        csrf_token = login_page.cookies.get("csrftoken", "")
        if not csrf_token:
            for line in login_page.text.splitlines():
                if "csrfmiddlewaretoken" in line:
                    parts = line.split('value="')
                    if len(parts) > 1:
                        csrf_token = parts[1].split('"')[0]
                        break
        session.post(
            f"{LS_URL}/user/login/",
            data={
                "email":               LS_USERNAME,
                "password":            LS_PASSWORD,
                "csrfmiddlewaretoken": csrf_token,
            },
            headers={
                "Referer":      f"{LS_URL}/user/login/",
                "Content-Type": "application/x-www-form-urlencoded",
            },
            allow_redirects=True,
            timeout=15,
        )
        token_resp = session.get(f"{LS_URL}/api/current-user/token", timeout=10)
        if token_resp.status_code == 200:
            token = token_resp.json().get("token", "")
            if token:
                print("[LS] ✓ API token fetched via session login.")
                return token
        print("[LS] Session login succeeded but no token returned.")
        return ""
    except Exception as e:
        print(f"[LS] Session login error: {e}")
        return ""


def _ls_get_working_headers() -> dict:
    """
    Return Authorization headers for the Label Studio API.
    Tries the manually-set token first, then the cached one,
    then attempts a fresh session login.
    """
    global _cached_ls_token, LS_API_TOKEN
    with _cached_token_lock:
        candidates = []
        if LS_API_TOKEN:
            candidates.append(LS_API_TOKEN)
        if _cached_ls_token and _cached_ls_token != LS_API_TOKEN:
            candidates.append(_cached_ls_token)

        for tok in candidates:
            headers = {"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}
            try:
                r = req_lib.get(f"{LS_URL}/api/current-user", headers=headers, timeout=5)
                if r.status_code == 200:
                    _cached_ls_token = tok
                    return headers
            except Exception:
                pass

        fresh = _ls_fetch_token_via_session()
        if fresh:
            _cached_ls_token = fresh
            return {"Authorization": f"Token {fresh}", "Content-Type": "application/json"}

        print("[LS] ✗ All auth methods failed.")
        return {}


def _ls_labeling_config() -> str:
    """Generate Label Studio polygon labeling XML configuration."""
    label_tags = "\n    ".join(
        f'<Label value="{cls}" background="{DEFECT_CLASS_COLORS.get(cls, "#888888")}"/>'
        for cls in DEFECT_CLASSES
    )
    return f"""<View>
  <Header value="Select label and click the image to start"/>
  <Image name="image" value="$image" zoom="true"/>

  <PolygonLabels name="label" toName="image" strokeWidth="3" pointSize="small" opacity="0.9">
    {label_tags}
  </PolygonLabels>
</View>"""


def _get_or_create_project(headers: dict = None) -> int | None:
    """Return the Label Studio project ID, creating the project if it doesn't exist."""
    if headers is None:
        headers = _ls_get_working_headers()
    if not headers:
        print("[LS] No working auth headers — cannot get/create project.")
        return None

    # Try cached project ID first.
    if os.path.exists(LS_PROJECT_ID_FILE):
        with open(LS_PROJECT_ID_FILE) as f:
            pid_str = f.read().strip()
        if pid_str.isdigit():
            pid = int(pid_str)
            try:
                r = req_lib.get(f"{LS_URL}/api/projects/{pid}", headers=headers, timeout=5)
                if r.status_code == 200:
                    return pid
                print(f"[LS] Cached project id={pid} not found — re-searching.")
            except Exception:
                pass

    # Search existing projects.
    try:
        r = req_lib.get(f"{LS_URL}/api/projects", headers=headers,
                        params={"page_size": 100}, timeout=10)
        if r.status_code == 200:
            data     = r.json()
            projects = data if isinstance(data, list) else data.get("results", [])
            for p in projects:
                if p.get("title") == PROJECT_TITLE:
                    pid = p["id"]
                    with open(LS_PROJECT_ID_FILE, "w") as f:
                        f.write(str(pid))
                    print(f"[LS] Found existing project '{PROJECT_TITLE}' id={pid}")
                    return pid
    except Exception as e:
        print(f"[LS] Error listing projects: {e}")

    # Create a new project.
    try:
        payload = {
            "title":        PROJECT_TITLE,
            "description":  "Draw polygons around blade defects.",
            "label_config": _ls_labeling_config(),
            "is_published": True,
            "sampling":     "Sequential",
        }
        r = req_lib.post(f"{LS_URL}/api/projects", headers=headers,
                         json=payload, timeout=15)
        if r.status_code in (200, 201):
            pid = r.json()["id"]
            with open(LS_PROJECT_ID_FILE, "w") as f:
                f.write(str(pid))
            print(f"[LS] Created project '{PROJECT_TITLE}' id={pid}")
            return pid
        print(f"[LS] Project creation failed {r.status_code}: {r.text[:300]}")
    except Exception:
        traceback.print_exc()

    return None


def _existing_task_filenames(pid: int, headers: dict) -> set:
    """Return the set of filenames already uploaded as tasks in a LS project."""
    names = set()
    page  = 1
    while True:
        try:
            r = req_lib.get(
                f"{LS_URL}/api/projects/{pid}/tasks",
                headers=headers,
                params={"page": page, "page_size": 100},
                timeout=15,
            )
            if r.status_code != 200:
                break
            data      = r.json()
            task_list = data if isinstance(data, list) else data.get("tasks", [])
            if not task_list:
                break
            for task in task_list:
                img_val = task.get("data", {}).get("image", "")
                names.add(os.path.basename(img_val.split("?d=")[-1]))
            if isinstance(data, list) or len(task_list) < 100:
                break
            page += 1
        except Exception:
            break
    return names


def _upload_image_as_task(pid: int, headers: dict, filepath: str, filename: str) -> bool:
    """Upload a single image file to Label Studio as a new task."""
    try:
        upload_headers = {"Authorization": headers["Authorization"]}
        with open(filepath, "rb") as f:
            r = req_lib.post(
                f"{LS_URL}/api/projects/{pid}/import",
                headers=upload_headers,
                files={"file": (filename, f, "image/png")},
                timeout=30,
            )
        if r.status_code in (200, 201):
            print(f"[LS Upload] ✓ '{filename}'")
            return True
        print(f"[LS Upload] ✗ '{filename}' — {r.status_code}: {r.text[:200]}")
        return False
    except Exception as e:
        print(f"[LS Upload] Exception for '{filename}': {e}")
        return False


def _import_new_files_to_ls(new_filenames: list):
    """Upload files from staging to Label Studio, skipping already-present ones."""
    try:
        headers = _ls_get_working_headers()
        if not headers:
            print("[LS Import] No working auth — skipping.")
            return
        pid = _get_or_create_project(headers)
        if not pid:
            print("[LS Import] No project — skipping.")
            return
        existing = _existing_task_filenames(pid, headers)
        to_add   = [fn for fn in new_filenames if fn not in existing]
        if not to_add:
            print("[LS Import] All files already in project.")
            return
        for fn in to_add:
            filepath = os.path.join(UPLOAD_STAGING, fn)
            if os.path.exists(filepath):
                _upload_image_as_task(pid, headers, filepath, fn)
    except Exception:
        traceback.print_exc()


def _full_sync_ls_project():
    """
    Wait for Label Studio to be healthy, then upload all staging images
    that are not yet present in the project.
    Designed to run in a daemon thread.
    """
    try:
        print("[LS Sync] Waiting for Label Studio to be ready...")
        for i in range(90):
            if _ls_is_running():
                print(f"[LS Sync] LS healthy after ~{i * 2}s")
                break
            time.sleep(2)
        else:
            print("[LS Sync] Timed out waiting for Label Studio.")
            return

        time.sleep(3)

        headers = _ls_get_working_headers()
        if not headers:
            print("[LS Sync] Could not authenticate — check credentials in backend.py")
            return

        print("[LS Sync] Authentication OK ✓")
        pid = _get_or_create_project(headers)
        if not pid:
            print("[LS Sync] Could not get/create project.")
            return

        print(f"[LS Sync] Project id={pid} ready.")
        staging_files = [
            f for f in os.listdir(UPLOAD_STAGING)
            if os.path.splitext(f)[1].lower() in ALLOWED_EXTENSIONS
        ]
        if not staging_files:
            print("[LS Sync] No staging images to upload.")
            return

        existing = _existing_task_filenames(pid, headers)
        to_add   = [f for f in staging_files if f not in existing]
        if not to_add:
            print("[LS Sync] All staging images already in Label Studio.")
            return

        print(f"[LS Sync] Uploading {len(to_add)} image(s)...")
        ok = fail = 0
        for fn in to_add:
            filepath = os.path.join(UPLOAD_STAGING, fn)
            if _upload_image_as_task(pid, headers, filepath, fn):
                ok += 1
            else:
                fail += 1
        print(f"[LS Sync] Done — {ok} uploaded, {fail} failed.")
    except Exception:
        traceback.print_exc()


def _ls_annotation_to_yolo(result: list, img_w: int, img_h: int) -> list:
    """Convert Label Studio annotation result items to YOLO polygon label lines."""
    lines = []
    for item in result:
        if item.get("type") not in ("rectanglelabels", "polygonlabels"):
            continue
        value  = item.get("value", {})
        labels = value.get("rectanglelabels", value.get("polygonlabels", []))
        if not labels:
            continue
        cls_idx = DEFECT_CLASS_INDEX.get(labels[0])
        if cls_idx is None:
            continue

        if item["type"] == "rectanglelabels":
            rx, ry = value["x"] / 100.0, value["y"] / 100.0
            rw, rh = value["width"] / 100.0, value["height"] / 100.0
            x1, y1 = rx,      ry
            x2, y2 = rx + rw, ry
            x3, y3 = rx + rw, ry + rh
            x4, y4 = rx,      ry + rh
            lines.append(
                f"{cls_idx} {x1:.6f} {y1:.6f} {x2:.6f} {y2:.6f} "
                f"{x3:.6f} {y3:.6f} {x4:.6f} {y4:.6f}"
            )

        elif item["type"] == "polygonlabels":
            points = value.get("points", [])
            if len(points) < 3:
                continue
            coords = []
            for px, py in points:
                coords.extend([px / 100.0, py / 100.0])
            lines.append(f"{cls_idx} " + " ".join(f"{c:.6f}" for c in coords))

    return lines


# =============================================================================
# ROUTES — SHOTS
# =============================================================================

@app.route('/shots', methods=['GET'])
def get_shots():
    try:
        return jsonify(_load_shots())
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/shots', methods=['POST'])
def save_shots():
    try:
        data = request.get_json(force=True)
        if not isinstance(data, list):
            return jsonify({"error": "Expected a JSON array of shots"}), 400
        with open(SHOTS_FILE, "w") as f:
            json.dump(data, f, indent=2)
        return jsonify({"message": f"Saved {len(data)} shots to shots.json"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# =============================================================================
# ROUTES — CAMERA
# =============================================================================

@app.route('/camera-start', methods=['POST'])
def camera_start():
    global _camera_process
    if _camera_process is not None and _camera_process.poll() is None:
        return jsonify({"status": "already_running", "pid": _camera_process.pid})
    _launch_camera_app()
    if _camera_process:
        return jsonify({"status": "started", "pid": _camera_process.pid})
    return jsonify({"error": "pic_clicker.py not found"}), 404


@app.route('/camera-stop', methods=['POST'])
def camera_stop():
    global _camera_process
    if _camera_process and _camera_process.poll() is None:
        _camera_process.terminate()
        try:
            _camera_process.wait(timeout=5)
        except subprocess.TimeoutExpired:
            _camera_process.kill()
        _camera_process = None
        return jsonify({"status": "stopped"})
    return jsonify({"status": "not_running"})


@app.route('/camera-stream')
def camera_stream():
    return Response(
        _generate_mjpeg(),
        mimetype='multipart/x-mixed-replace; boundary=frame',
    )


@app.route('/camera-snap')
def camera_snap():
    with _camera_lock:
        frame = _camera_frame
    if frame is None:
        return jsonify({
            "error": "No camera frame available — open the camera in the control panel first."
        }), 503
    return Response(frame, mimetype='image/jpeg')


@app.route('/camera-status')
def camera_status():
    with _camera_lock:
        available = _camera_frame is not None
        meta      = dict(_camera_meta)
    return jsonify({"available": available, **meta})


# =============================================================================
# ROUTES — PIC CLICKER / CAPTURE SESSION
# =============================================================================

@app.route('/start-capture', methods=['POST'])
def start_capture():
    global _pic_clicker_process, _pic_clicker_folder
    try:
        data        = request.get_json() or {}
        folder_name = data.get('folder_name', '').strip()
        if not folder_name:
            return jsonify({"error": "Folder name is required"}), 400

        with _pic_clicker_lock:
            if _pic_clicker_process is not None:
                try:
                    _pic_clicker_process.terminate()
                    _pic_clicker_process.wait(timeout=3)
                except Exception:
                    _pic_clicker_process.kill()
                _pic_clicker_process = None

            capture_folder   = os.path.join(BASE_DIR, folder_name)
            os.makedirs(capture_folder, exist_ok=True)

            pic_clicker_path = os.path.join(BASE_DIR, "pic_clicker.py")
            if not os.path.exists(pic_clicker_path):
                return jsonify({"error": f"pic_clicker.py not found at {pic_clicker_path}"}), 404

            _pic_clicker_process = subprocess.Popen(
                [sys.executable, pic_clicker_path, capture_folder], cwd=BASE_DIR
            )
            _pic_clicker_folder = folder_name

        return jsonify({
            "message": "Pic clicker started",
            "folder":  folder_name,
            "pid":     _pic_clicker_process.pid,
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/stop-capture', methods=['POST'])
def stop_capture():
    global _pic_clicker_process
    try:
        with _pic_clicker_lock:
            if _pic_clicker_process is not None:
                try:
                    _pic_clicker_process.terminate()
                    _pic_clicker_process.wait(timeout=3)
                except Exception:
                    _pic_clicker_process.kill()
                _pic_clicker_process = None
        return jsonify({"message": "Pic clicker stopped"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/capture-status', methods=['GET'])
def capture_status():
    process_running = (
        _pic_clicker_process is not None and _pic_clicker_process.poll() is None
    )
    return jsonify({
        "running": process_running,
        "folder":  _pic_clicker_folder,
        "pid":     _pic_clicker_process.pid if _pic_clicker_process else None,
    })


@app.route('/list-captured-images', methods=['GET'])
def list_captured_images():
    try:
        folder_name = request.args.get('folder', _pic_clicker_folder)
        if not folder_name:
            return jsonify({"error": "Folder name required"}), 400

        capture_folder = os.path.join(BASE_DIR, folder_name)
        if not os.path.isdir(capture_folder):
            return jsonify({"images": [], "count": 0})

        images = sorted([
            f for f in os.listdir(capture_folder)
            if os.path.splitext(f)[1].lower() in ALLOWED_EXTENSIONS
        ])
        return jsonify({
            "folder": folder_name,
            "images": [
                {"filename": f, "url": f"http://localhost:5000/captured-image/{folder_name}/{f}"}
                for f in images
            ],
            "count": len(images),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/captured-image/<folder_name>/<filename>', methods=['GET'])
def serve_captured_image(folder_name, filename):
    try:
        path = os.path.join(os.path.abspath(BASE_DIR), folder_name, filename)
        if not os.path.abspath(path).startswith(os.path.abspath(BASE_DIR)):
            return jsonify({"error": "Invalid path"}), 403
        if not os.path.exists(path):
            return jsonify({"error": "File not found"}), 404
        return send_file(path)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# =============================================================================
# ROUTES — PREDICTION
# =============================================================================

@app.route('/predict', methods=['POST'])
def predict_route():
    try:
        if 'image' not in request.files:
            return jsonify({"error": "No image file provided"}), 400

        file  = request.files['image']
        image = cv2.imdecode(np.frombuffer(file.read(), np.uint8), cv2.IMREAD_COLOR)
        if image is None:
            return jsonify({"error": "Failed to decode uploaded image"}), 400

        result = _process_image_pipeline(image, filename=file.filename or "image")
        if result is None:
            return jsonify({"error": "No body detected in image"}), 400

        excel_name = f"report_{uuid.uuid4().hex[:8]}.xlsx"
        _save_excel(
            [{
                "filename":        result["filename"],
                "body_confidence": result["body_confidence"],
                "detections":      result["detections"],
                "blade_label":     None,
            }],
            os.path.join(OUTPUT_DIR, excel_name),
        )

        return jsonify({
            "predicted_image": _encode_image_b64(result["predicted_image"]),
            "detections":      result["detections"],
            "excel_report":    excel_name,
            "body_message":    result.get("body_message"),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/predict-batch', methods=['POST'])
def predict_batch():
    try:
        files = request.files.getlist("images")
        if not files:
            return jsonify({"error": "No images provided"}), 400

        batch_results, excel_rows = [], []

        for file in files:
            image  = cv2.imdecode(np.frombuffer(file.read(), np.uint8), cv2.IMREAD_COLOR)
            result = _process_image_pipeline(image, filename=file.filename or "image")

            if result is None:
                batch_results.append({"filename": file.filename, "error": "No body detected", "detections": []})
                excel_rows.append({
                    "filename": file.filename, "body_confidence": 0.0,
                    "detections": [], "blade_label": None,
                })
                continue

            batch_results.append({
                "filename":        result["filename"],
                "predicted_image": _encode_image_b64(result["predicted_image"]),
                "detections":      result["detections"],
                "body_message":    result.get("body_message"),
            })
            excel_rows.append({
                "filename":        result["filename"],
                "body_confidence": result["body_confidence"],
                "detections":      result["detections"],
                "blade_label":     None,
            })

        excel_name = f"batch_report_{uuid.uuid4().hex[:8]}.xlsx"
        _save_excel(excel_rows, os.path.join(OUTPUT_DIR, excel_name))

        return jsonify({"results": batch_results, "excel_report": excel_name})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/process-captured-images', methods=['POST'])
def process_captured_images():
    try:
        data        = request.get_json() or {}
        folder_name = data.get('folder_name', '').strip() or _pic_clicker_folder
        filenames   = data.get('filenames')

        if not folder_name:
            return jsonify({"error": "folder_name is required (or start a capture session first)"}), 400

        capture_folder = os.path.join(BASE_DIR, folder_name)
        if not os.path.isdir(capture_folder):
            return jsonify({"error": f"Folder not found: {folder_name}"}), 404

        candidates = (
            [f for f in filenames if os.path.splitext(f)[1].lower() in ALLOWED_EXTENSIONS]
            if filenames
            else sorted([
                f for f in os.listdir(capture_folder)
                if os.path.splitext(f)[1].lower() in ALLOWED_EXTENSIONS
            ])
        )
        if not candidates:
            return jsonify({"error": "No valid images found in folder"}), 400

        batch_results, excel_rows = [], []

        for filename in candidates:
            filepath = os.path.join(capture_folder, filename)
            if not os.path.exists(filepath):
                batch_results.append({"filename": filename, "error": "File not found"})
                continue

            image = cv2.imread(filepath)
            if image is None:
                batch_results.append({"filename": filename, "error": "Failed to decode image"})
                continue

            result = _process_image_pipeline(image, filename=filename)

            if result is None:
                batch_results.append({"filename": filename, "error": "No body detected", "detections": []})
                excel_rows.append({
                    "filename": filename, "body_confidence": 0.0,
                    "detections": [], "blade_label": None,
                })
                continue

            batch_results.append({
                "filename":        result["filename"],
                "predicted_image": _encode_image_b64(result["predicted_image"]),
                "detections":      result["detections"],
                "body_message":    result.get("body_message"),
            })
            excel_rows.append({
                "filename":        result["filename"],
                "body_confidence": result["body_confidence"],
                "detections":      result["detections"],
                "blade_label":     None,
            })

        excel_name = f"capture_report_{uuid.uuid4().hex[:8]}.xlsx"
        if excel_rows:
            _save_excel(excel_rows, os.path.join(OUTPUT_DIR, excel_name))

        return jsonify({
            "results":      batch_results,
            "excel_report": excel_name if excel_rows else None,
            "folder":       folder_name,
            "total":        len(candidates),
            "processed":    len(excel_rows),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/download/<filename>', methods=['GET'])
def download_excel(filename):
    path = os.path.join(OUTPUT_DIR, filename)
    if not os.path.exists(path):
        return jsonify({"error": "File not found"}), 404
    return send_file(
        path, as_attachment=True, download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


# =============================================================================
# ROUTES — RETRAIN DATA MANAGEMENT
# =============================================================================

@app.route('/save-label', methods=['POST'])
def save_label():
    try:
        data       = request.get_json(force=True)
        filename   = data.get("filename", f"img_{uuid.uuid4().hex[:8]}.png")
        label      = data.get("label")
        image_b64  = data.get("image_b64")
        detections = data.get("detections", [])

        if not label or not image_b64:
            return jsonify({"error": "Missing label or image_b64"}), 400
        if label not in ("approved", "defective", "retrain"):
            return jsonify({"error": "label must be approved/defective/retrain"}), 400

        safe     = os.path.basename(secure_filename(filename))
        base     = os.path.splitext(safe)[0]
        img_name = base + ".png"

        target_img, target_lbl = {
            "approved":  (APPROVED_DIR,      APPROVED_LBL),
            "defective": (DEFECTIVE_DIR,      DEFECTIVE_LBL),
            "retrain":   (RETRAIN_QUEUE_DIR,  RETRAIN_QUEUE_LBL),
        }[label]

        img_dest  = os.path.join(target_img, img_name)
        img_bytes = base64.b64decode(image_b64)
        with open(img_dest, "wb") as f:
            f.write(img_bytes)

        lbl_path = os.path.join(target_lbl, base + ".txt")
        _write_yolo_label(detections, img_dest, lbl_path)

        if detections:
            with open(os.path.join(target_lbl, base + ".json"), "w") as f:
                json.dump({
                    "filename":   img_name,
                    "label":      label,
                    "detections": detections,
                    "saved_at":   datetime.now().isoformat(),
                }, f, indent=2)

        _append_log(img_name, label)
        return jsonify({"message": f"Saved '{img_name}' as '{label}'"})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/retrain-stats', methods=['GET'])
def retrain_stats():
    history = []
    if os.path.exists(RETRAIN_LOG):
        try:
            with open(RETRAIN_LOG) as f:
                history = json.load(f)
        except Exception:
            history = []

    approved  = _count_images_in_dir(APPROVED_DIR)
    defective = _count_images_in_dir(DEFECTIVE_DIR)
    retrain   = _count_images_in_dir(RETRAIN_QUEUE_DIR)
    staging   = _count_images_in_dir(UPLOAD_STAGING)

    label_stats = {
        "approved":  _audit_trainable(APPROVED_DIR,      APPROVED_LBL),
        "defective": _audit_trainable(DEFECTIVE_DIR,      DEFECTIVE_LBL),
        "retrain":   _audit_trainable(RETRAIN_QUEUE_DIR,  RETRAIN_QUEUE_LBL),
    }

    return jsonify({
        "approved":    approved,
        "defective":   defective,
        "retrain":     retrain,
        "total":       approved + defective + retrain,
        "staging":     staging,
        "history":     history[-20:],
        "label_stats": label_stats,
    })


@app.route('/import-yolo-zip', methods=['POST'])
def import_yolo_zip():
    if 'zip' not in request.files:
        return jsonify({"error": "No zip file provided — send it as form-data field 'zip'"}), 400

    zip_file = request.files['zip']
    if not zip_file.filename.lower().endswith('.zip'):
        return jsonify({"error": "Uploaded file must be a .zip"}), 400

    tmp_dir = os.path.join(RETRAIN_DIR, f"zip_import_{uuid.uuid4().hex[:8]}")
    os.makedirs(tmp_dir, exist_ok=True)

    try:
        zip_path = os.path.join(tmp_dir, "upload.zip")
        zip_file.save(zip_path)

        extract_dir = os.path.join(tmp_dir, "extracted")
        os.makedirs(extract_dir, exist_ok=True)
        try:
            with zipfile.ZipFile(zip_path, 'r') as zf:
                zf.extractall(extract_dir)
        except zipfile.BadZipFile:
            return jsonify({"error": "Uploaded file is not a valid zip archive"}), 400

        image_files: dict = {}
        label_files: dict = {}

        for root, dirs, files in os.walk(extract_dir):
            dirs[:] = [d for d in dirs if not d.startswith('__') and not d.startswith('.')]
            for fn in files:
                if fn.startswith('.'):
                    continue
                ext  = os.path.splitext(fn)[1].lower()
                stem = os.path.splitext(fn)[0]
                full = os.path.join(root, fn)
                if ext in ALLOWED_EXTENSIONS_WIDE:
                    if stem not in image_files or "image" in root.lower():
                        image_files[stem] = full
                elif ext == '.txt' and fn.lower() != 'classes.txt':
                    if stem not in label_files or "label" in root.lower():
                        label_files[stem] = full

        all_stems        = set(image_files) | set(label_files)
        unmatched_images = sorted(
            os.path.basename(v) for k, v in image_files.items() if k not in label_files
        )
        unmatched_labels = sorted(
            os.path.basename(v) for k, v in label_files.items() if k not in image_files
        )

        preview = []
        for stem in sorted(all_stems):
            src_img = image_files.get(stem)
            src_lbl = label_files.get(stem)
            if src_img is None:
                continue

            orig_ext  = os.path.splitext(src_img)[1].lower()
            safe_stem = secure_filename(stem) or uuid.uuid4().hex[:8]
            img_name  = safe_stem + orig_ext
            dst_img   = os.path.join(DEFECTIVE_DIR, img_name)
            dst_lbl   = os.path.join(DEFECTIVE_LBL, safe_stem + ".txt")

            # Resolve filename collision, keeping image and label stems in sync.
            if os.path.exists(dst_img):
                safe_stem = f"{safe_stem}_{uuid.uuid4().hex[:6]}"
                img_name  = safe_stem + orig_ext
                dst_img   = os.path.join(DEFECTIVE_DIR, img_name)
                dst_lbl   = os.path.join(DEFECTIVE_LBL, safe_stem + ".txt")

            shutil.copy2(src_img, dst_img)

            label_count = 0
            if src_lbl and os.path.exists(src_lbl):
                try:
                    with open(src_lbl) as lf:
                        lbl_content = lf.read().strip()
                    label_count = sum(1 for line in lbl_content.splitlines() if line.strip())
                except Exception:
                    lbl_content = ""
                with open(dst_lbl, "w") as lf:
                    lf.write(lbl_content)
            else:
                open(dst_lbl, "w").close()

            _append_log(img_name, "defective", source="zip_import")
            preview.append({"filename": img_name, "label_count": label_count})

        matched = sum(1 for p in preview if p["label_count"] > 0)

        return jsonify({
            "images_found":     len(image_files),
            "labels_found":     len(label_files),
            "matched":          matched,
            "unmatched_images": unmatched_images,
            "unmatched_labels": unmatched_labels,
            "preview":          preview,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)


@app.route('/retrain', methods=['POST'])
def retrain():
    global _retrain_process

    with _retrain_lock:
        if _retrain_process is not None and _retrain_process.poll() is None:
            return jsonify({"error": "Training already in progress. Stop it first."}), 409

        try:
            body       = request.get_json(force=True) if request.content_length else {}
            epochs     = int(body.get("epochs", 50))
            batch_size = int(body.get("batch_size", 4))
            lr         = float(body.get("lr", 3e-4))

            script_path   = os.path.join(BASE_DIR, "BFL_retrain.py")
            notebook_path = os.path.join(BASE_DIR, "BFL_retrain.ipynb")

            if not os.path.exists(script_path) and not os.path.exists(notebook_path):
                return jsonify({"error": "Neither BFL_retrain.py nor BFL_retrain.ipynb found."}), 404

            total_annotated = (
                _audit_trainable(DEFECTIVE_DIR,     DEFECTIVE_LBL)["with_labels"] +
                _audit_trainable(RETRAIN_QUEUE_DIR, RETRAIN_QUEUE_LBL)["with_labels"]
            )
            if total_annotated == 0:
                return jsonify({
                    "error": (
                        "No annotated images found. Each image needs a matching "
                        ".txt label file with at least one polygon. "
                        "Import a YOLO zip via the Import Data tab."
                    )
                }), 400

            for path in [RETRAIN_STOP_FILE, RETRAIN_PROGRESS_FILE]:
                if os.path.exists(path):
                    os.remove(path)

            env = os.environ.copy()
            env.update({
                "RETRAIN_APPROVED_DIR":     os.path.abspath(APPROVED_DIR),
                "RETRAIN_DEFECTIVE_DIR":    os.path.abspath(DEFECTIVE_DIR),
                "RETRAIN_QUEUE_DIR":        os.path.abspath(RETRAIN_QUEUE_DIR),
                "RETRAIN_OUTPUT_MODEL":     os.path.abspath("hr_net_retrained.pth"),
                "RETRAIN_BASE_MODEL":       os.path.abspath("hr_net.pth"),
                "RETRAIN_EPOCHS":           str(epochs),
                "RETRAIN_BATCH_SIZE":       str(batch_size),
                "RETRAIN_LR":               str(lr),
                "RETRAIN_PROGRESS_FILE":    os.path.abspath(RETRAIN_PROGRESS_FILE),
                "RETRAIN_STOP_FILE":        os.path.abspath(RETRAIN_STOP_FILE),
                "RETRAIN_CLASS_NAMES_FILE": os.path.join(RETRAIN_DIR, "class_names.json"),
            })

            if os.path.exists(script_path):
                cmd = [sys.executable, script_path]
            else:
                executed_nb = f"BFL_executed_{datetime.now().strftime('%Y%m%d_%H%M%S')}.ipynb"
                cmd = [
                    sys.executable, "-m", "nbconvert", "--to", "notebook", "--execute",
                    "--ExecutePreprocessor.timeout=7200",
                    "--output", executed_nb, notebook_path,
                ]

            _start_retrain_process(cmd, env)
            print(
                f"[Retrain] Started PID={_retrain_process.pid} "
                f"epochs={epochs} batch={batch_size} lr={lr} annotated={total_annotated}"
            )

            return jsonify({
                "message":   "Training started",
                "pid":       _retrain_process.pid,
                "epochs":    epochs,
                "annotated": total_annotated,
            })

        except Exception as e:
            traceback.print_exc()
            return jsonify({"error": str(e)}), 500


@app.route('/retrain-stop', methods=['POST'])
def retrain_stop():
    global _retrain_process
    try:
        with open(RETRAIN_STOP_FILE, "w") as f:
            f.write("stop")
    except Exception as e:
        return jsonify({"error": f"Could not write stop signal: {e}"}), 500

    with _retrain_lock:
        if _retrain_process is not None and _retrain_process.poll() is None:
            for _ in range(10):
                time.sleep(1)
                if _retrain_process.poll() is not None:
                    break
            else:
                _retrain_process.terminate()
                try:
                    _retrain_process.wait(timeout=5)
                except subprocess.TimeoutExpired:
                    _retrain_process.kill()
            _retrain_process = None

    return jsonify({"message": "Stop signal sent."})


@app.route('/retrain-progress', methods=['GET'])
def retrain_progress():
    process_running = _retrain_process is not None and _retrain_process.poll() is None

    if not os.path.exists(RETRAIN_PROGRESS_FILE):
        return jsonify({
            "status":          "training" if process_running else "idle",
            "epoch":           0,
            "total_epochs":    0,
            "train_loss":      None,
            "val_loss":        None,
            "best_val":        None,
            "message":         "Waiting for training to start…" if process_running else "No training in progress.",
            "history":         [],
            "process_running": process_running,
            "retrain_used":    0,
            "defective_used":  0,
        })

    try:
        with open(RETRAIN_PROGRESS_FILE) as f:
            data = json.load(f)

        data["process_running"] = process_running
        data.setdefault("retrain_used",   0)
        data.setdefault("defective_used", 0)

        if not process_running and data.get("status") == "training":
            rc             = _retrain_process.returncode if _retrain_process is not None else 0
            data["status"] = "done" if rc == 0 else "error"
            data["message"] = (
                f"Training process ended (exit code {rc})."
                if rc != 0
                else data.get("message", "Training complete.")
            )

        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e), "process_running": process_running}), 500


@app.route('/retrain-log', methods=['GET'])
def retrain_log():
    """Return the last N lines from the training stdout log."""
    try:
        n = int(request.args.get('lines', 80))
    except Exception:
        n = 80

    running = _retrain_process is not None and _retrain_process.poll() is None

    if not os.path.exists(RETRAIN_LOG_FILE):
        return jsonify({"lines": [], "process_running": running, "exists": False})

    try:
        with open(RETRAIN_LOG_FILE, 'r', errors='replace') as f:
            all_lines = f.readlines()
        return jsonify({
            "lines":           [l.rstrip('\n') for l in all_lines[-n:]],
            "total_lines":     len(all_lines),
            "process_running": running,
            "exists":          True,
        })
    except Exception as e:
        return jsonify({"error": str(e), "lines": [], "exists": False}), 500


@app.route('/retrain-debug', methods=['GET'])
def retrain_debug():
    """Full audit of every image/label pair — useful for diagnosing training-gate issues."""
    def audit(img_dir, lbl_dir):
        results = []
        if not os.path.isdir(img_dir):
            return results
        for fn in sorted(os.listdir(img_dir)):
            if os.path.splitext(fn)[1].lower() not in ALLOWED_EXTENSIONS:
                continue
            lbl_path = os.path.join(lbl_dir, os.path.splitext(fn)[0] + ".txt")
            exists   = os.path.exists(lbl_path)
            size = lines = 0
            if exists:
                try:
                    size = os.path.getsize(lbl_path)
                    with open(lbl_path) as f:
                        lines = sum(1 for l in f if l.strip())
                except Exception:
                    pass
            results.append({
                "image":         fn,
                "label_path":    lbl_path,
                "label_exists":  exists,
                "label_size":    size,
                "polygon_count": lines,
                "trainable":     exists and size > 0,
            })
        return results

    defective_audit = audit(DEFECTIVE_DIR,     DEFECTIVE_LBL)
    retrain_audit   = audit(RETRAIN_QUEUE_DIR, RETRAIN_QUEUE_LBL)

    return jsonify({
        "total_annotated_for_training": (
            sum(1 for r in defective_audit if r["trainable"]) +
            sum(1 for r in retrain_audit   if r["trainable"])
        ),
        "defective": defective_audit,
        "retrain":   retrain_audit,
    })


@app.route('/retrain-clear', methods=['POST'])
def retrain_clear():
    global _retrain_process
    with _retrain_lock:
        if _retrain_process is not None and _retrain_process.poll() is None:
            return jsonify({"error": "Training is currently running. Stop it first."}), 409
    try:
        for d in [APPROVED_DIR, DEFECTIVE_DIR, APPROVED_LBL, DEFECTIVE_LBL,
                  RETRAIN_QUEUE_DIR, RETRAIN_QUEUE_LBL]:
            shutil.rmtree(d, ignore_errors=True)
            os.makedirs(d, exist_ok=True)
        for path in [RETRAIN_LOG, RETRAIN_PROGRESS_FILE, RETRAIN_STOP_FILE]:
            if os.path.exists(path):
                os.remove(path)
        return jsonify({"message": "Retrain data cleared"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# =============================================================================
# ROUTES — STAGING
# =============================================================================

@app.route('/upload-new-images', methods=['POST'])
def upload_new_images():
    try:
        files = request.files.getlist("images")
        if not files:
            return jsonify({"error": "No images provided"}), 400

        saved, skipped = [], []
        for file in files:
            ext = os.path.splitext(file.filename)[1].lower()
            if ext not in ALLOWED_EXTENSIONS:
                skipped.append(file.filename)
                continue
            safe      = secure_filename(file.filename)
            base, e   = os.path.splitext(safe)
            dest      = os.path.join(UPLOAD_STAGING, safe)
            if os.path.exists(dest):
                safe = f"{base}_{uuid.uuid4().hex[:6]}{e}"
                dest = os.path.join(UPLOAD_STAGING, safe)
            file.save(dest)
            saved.append(safe)

        ls_running = _ls_is_running()
        if ls_running and saved:
            threading.Thread(target=_import_new_files_to_ls, args=(saved,), daemon=True).start()

        return jsonify({
            "saved":         saved,
            "skipped":       skipped,
            "staging_total": _count_staging(),
            "ls_synced":     ls_running,
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/staging-images', methods=['GET'])
def list_staging_images():
    images = sorted([
        f for f in os.listdir(UPLOAD_STAGING)
        if os.path.splitext(f)[1].lower() in ALLOWED_EXTENSIONS
    ])
    return jsonify({
        "images": [
            {"filename": f, "url": f"http://localhost:5000/staging-preview/{f}"}
            for f in images
        ],
        "count": len(images),
    })


@app.route('/staging-preview/<filename>', methods=['GET'])
def serve_staging_image(filename):
    path = os.path.join(os.path.abspath(UPLOAD_STAGING), secure_filename(filename))
    if not os.path.exists(path):
        return jsonify({"error": "Not found"}), 404
    return send_file(path)


@app.route('/staging-images/<filename>', methods=['DELETE'])
def delete_staging_image(filename):
    path = os.path.join(UPLOAD_STAGING, secure_filename(filename))
    if os.path.exists(path):
        os.remove(path)
    lbl = os.path.join(
        UPLOAD_ANNOT_LBL,
        os.path.splitext(secure_filename(filename))[0] + ".txt",
    )
    if os.path.exists(lbl):
        os.remove(lbl)
    return jsonify({"message": f"Deleted {filename}"})


@app.route('/promote-staging', methods=['POST'])
def promote_staging():
    try:
        data        = request.get_json(force=True)
        assignments = data.get("assignments", [])
        if not assignments:
            return jsonify({"error": "No assignments provided"}), 400

        promoted, errors, labels_found = [], [], 0

        for item in assignments:
            fn    = secure_filename(item.get("filename", ""))
            label = item.get("label")
            if not fn or label not in ("approved", "defective", "retrain"):
                errors.append({"filename": fn, "error": "Invalid entry"})
                continue

            src = os.path.join(UPLOAD_STAGING, fn)
            if not os.path.exists(src):
                errors.append({"filename": fn, "error": "Not in staging"})
                continue

            dst_img, dst_lbl = {
                "approved":  (APPROVED_DIR,      APPROVED_LBL),
                "defective": (DEFECTIVE_DIR,      DEFECTIVE_LBL),
                "retrain":   (RETRAIN_QUEUE_DIR,  RETRAIN_QUEUE_LBL),
            }[label]

            shutil.move(src, os.path.join(dst_img, fn))
            base          = os.path.splitext(fn)[0]
            lbl_src_txt   = os.path.join(UPLOAD_ANNOT_LBL, base + ".txt")
            lbl_src_json  = os.path.join(UPLOAD_ANNOT_LBL, base + ".json")

            if os.path.exists(lbl_src_txt):
                shutil.move(lbl_src_txt, os.path.join(dst_lbl, base + ".txt"))
                labels_found += 1
            elif os.path.exists(lbl_src_json):
                try:
                    with open(lbl_src_json) as jf:
                        jdata = json.load(jf)
                    dets = jdata if isinstance(jdata, list) else jdata.get("detections", [])
                    _write_yolo_label(dets, os.path.join(dst_img, fn),
                                      os.path.join(dst_lbl, base + ".txt"))
                    os.remove(lbl_src_json)
                    labels_found += 1
                except Exception:
                    open(os.path.join(dst_lbl, base + ".txt"), "w").close()
            else:
                open(os.path.join(dst_lbl, base + ".txt"), "w").close()

            _append_log(fn, label, source="staging")
            promoted.append(fn)

        return jsonify({
            "promoted":          promoted,
            "labels_found":      labels_found,
            "errors":            errors,
            "staging_remaining": _count_staging(),
        })
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# =============================================================================
# ROUTES — LABEL STUDIO / ANNOTATION SERVER
# =============================================================================

@app.route('/annotation-server/start', methods=['POST'])
def start_annotation_server():
    global _ls_process
    with _ls_lock:
        if _ls_is_running():
            threading.Thread(target=_full_sync_ls_project, daemon=True).start()
            return jsonify({
                "status": "running",
                "url":    LS_URL,
                "note":   "Already running — syncing images.",
            })

        env = os.environ.copy()
        env.update({
            "LABEL_STUDIO_LOCAL_FILES_SERVING_ENABLED": "true",
            "LABEL_STUDIO_LOCAL_FILES_DOCUMENT_ROOT":   os.path.abspath(RETRAIN_DIR),
            "LABEL_STUDIO_USERNAME":                    LS_USERNAME,
            "LABEL_STUDIO_PASSWORD":                    LS_PASSWORD,
            "LABEL_STUDIO_DISABLE_SIGNUP_WITHOUT_LINK": "false",
        })

        launch_cmds = [
            ["label-studio", "start", "--port", str(LS_PORT),
             "--data-dir", os.path.abspath(LS_DATA_DIR), "--no-browser"],
            ["python", "-m", "label_studio", "start", "--port", str(LS_PORT),
             "--data-dir", os.path.abspath(LS_DATA_DIR), "--no-browser"],
        ]

        launched = False
        for cmd in launch_cmds:
            try:
                _ls_process = subprocess.Popen(
                    cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, env=env,
                )
                launched = True
                print(f"[LS] Launched: {' '.join(cmd)}")
                break
            except FileNotFoundError:
                continue

        if not launched:
            return jsonify({"error": "label-studio not found. Run: pip install label-studio"}), 500

    threading.Thread(target=_full_sync_ls_project, daemon=True).start()
    return jsonify({
        "status": "starting",
        "url":    LS_URL,
        "note":   "Label Studio is starting. Project and polygon labels will be created automatically.",
    })


@app.route('/annotation-server/stop', methods=['POST'])
def stop_annotation_server():
    global _ls_process
    with _ls_lock:
        if _ls_process and _ls_process.poll() is None:
            _ls_process.terminate()
            try:
                _ls_process.wait(timeout=5)
            except subprocess.TimeoutExpired:
                _ls_process.kill()
            _ls_process = None
            return jsonify({"status": "stopped"})
    return jsonify({"status": "not_running"})


@app.route('/annotation-server/status', methods=['GET'])
def annotation_server_status():
    running  = _ls_is_running()
    token_ok = bool(_ls_get_working_headers()) if running else False
    return jsonify({
        "status":    "running" if running else (
                     "starting" if _ls_process and _ls_process.poll() is None
                     else "stopped"),
        "url":       LS_URL,
        "token_set": bool(LS_API_TOKEN or _cached_ls_token),
        "token_ok":  token_ok,
    })


@app.route('/annotation-server/sync', methods=['POST'])
def sync_annotation_server():
    if not _ls_is_running():
        return jsonify({"error": "Label Studio is not running"}), 400
    threading.Thread(target=_full_sync_ls_project, daemon=True).start()
    return jsonify({"message": "Sync started — images uploading to Label Studio."})


@app.route('/annotation-server/setup', methods=['POST'])
def force_setup():
    threading.Thread(target=_full_sync_ls_project, daemon=True).start()
    return jsonify({"message": "Setup triggered. Watch backend terminal for progress."})


@app.route('/annotation-server/token', methods=['POST'])
def set_ls_token():
    global LS_API_TOKEN, _cached_ls_token
    data  = request.get_json(force=True)
    token = data.get("token", "").strip()
    if not token:
        return jsonify({"error": "No token provided"}), 400

    LS_API_TOKEN     = token
    _cached_ls_token = token

    if not _ls_is_running():
        return jsonify({"message": "Token saved. Start Label Studio to validate it."})

    headers = _ls_get_working_headers()
    if headers:
        return jsonify({"message": "Token saved and validated ✓", "valid": True})
    return jsonify({"message": "Token saved but validation failed — check it.", "valid": False})


@app.route('/annotation-server/debug', methods=['GET'])
def debug_ls():
    running = _ls_is_running()
    result  = {
        "ls_running":    running,
        "ls_url":        LS_URL,
        "token_cached":  bool(_cached_ls_token),
        "token_manual":  bool(LS_API_TOKEN),
        "staging_count": _count_staging(),
    }
    if not running:
        result["note"] = "Label Studio is not running."
        return jsonify(result)

    working_headers = _ls_get_working_headers()
    if working_headers:
        pid = _get_or_create_project(working_headers)
        result["project_id"] = pid
        result["auth_ok"]    = True
        if pid:
            result["tasks_in_ls"] = len(_existing_task_filenames(pid, working_headers))
    else:
        result["auth_ok"] = False
        result["note"]    = "All auth methods failed. Try POST /annotation-server/token with your JWT."

    return jsonify(result)


@app.route('/annotation-server/pull-annotations', methods=['POST'])
def pull_annotations():
    if not _ls_is_running():
        return jsonify({"error": "Label Studio is not running"}), 400

    headers = _ls_get_working_headers()
    if not headers:
        return jsonify({"error": "Label Studio auth failed"}), 500

    pid = _get_or_create_project(headers)
    if not pid:
        return jsonify({"error": "No project found"}), 500

    pulled, skipped = [], []
    page = 1
    while True:
        try:
            r = req_lib.get(
                f"{LS_URL}/api/projects/{pid}/tasks",
                headers=headers,
                params={"page": page, "page_size": 100},
                timeout=30,
            )
            if r.status_code != 200:
                break
            data      = r.json()
            task_list = data if isinstance(data, list) else data.get("tasks", [])
            if not task_list:
                break

            for task in task_list:
                annotations = task.get("annotations", [])
                if not annotations:
                    skipped.append(task.get("id"))
                    continue

                img_val      = task.get("data", {}).get("image", "")
                img_name     = os.path.basename(img_val.split("?d=")[-1])
                base         = os.path.splitext(img_name)[0]
                staging_path = os.path.join(UPLOAD_STAGING, img_name)

                img_w = img_h = 1
                if os.path.exists(staging_path):
                    img = cv2.imread(staging_path)
                    if img is not None:
                        img_h, img_w = img.shape[:2]

                lines    = _ls_annotation_to_yolo(annotations[0].get("result", []), img_w, img_h)
                lbl_path = os.path.join(UPLOAD_ANNOT_LBL, base + ".txt")
                with open(lbl_path, "w") as lf:
                    lf.write("\n".join(lines))
                pulled.append(img_name)

            if isinstance(data, list) or len(task_list) < 100:
                break
            page += 1
        except Exception:
            traceback.print_exc()
            break

    return jsonify({
        "pulled":         len(pulled),
        "skipped_no_ann": len(skipped),
        "files":          pulled,
    })


@app.route('/annotation-server/upload-annotations', methods=['POST'])
def upload_annotations():
    try:
        files = request.files.getlist("labels")
        if not files:
            return jsonify({"error": "No files provided"}), 400

        saved, skipped = [], []
        for file in files:
            if not file.filename.lower().endswith(".txt"):
                skipped.append(file.filename)
                continue
            safe = secure_filename(file.filename)
            file.save(os.path.join(UPLOAD_ANNOT_LBL, safe))
            saved.append(safe)

        return jsonify({"saved": saved, "skipped": skipped})
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


# =============================================================================
# MAIN
# =============================================================================

if __name__ == '__main__':
    _ensure_directories()
    _clear_stale_files()
    _load_shots()

    print("Initializing models...")
    body_model.initialize_model("body_model.pt")
    hrnet_model.initialize_model("hr_net.pth")
    print("All models loaded.")

    print("Starting Flask server on port 5000...")
    app.run(debug=False, port=5000)