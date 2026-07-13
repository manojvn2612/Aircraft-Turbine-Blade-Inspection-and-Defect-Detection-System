import sys, uvcham
import os
from datetime import datetime
from PyQt5.QtCore import pyqtSignal, pyqtSlot, QTimer, QSignalBlocker, Qt
from PyQt5.QtGui import QPixmap, QImage
from PyQt5.QtWidgets import (QLabel, QApplication, QWidget, QCheckBox,
    QMessageBox, QPushButton, QComboBox, QSlider, QGroupBox, QGridLayout,
    QBoxLayout, QHBoxLayout, QVBoxLayout, QMenu, QAction, QLineEdit,
    QInputDialog)
import numpy as np
import cv2
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

XLSX_HEADERS = ["SR", "Part Name", "Elevation", "Cam angle", "Table angle", "Zoom", "Focus", "Flash"]
XLSX_FILENAME = "Blade_angles.xlsx"


def get_xlsx_path():
    # print("path", os.path.join(os.path.dirname(os.path.abspath(__file__)), XLSX_FILENAME))
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), XLSX_FILENAME)


def ensure_xlsx():
    path = get_xlsx_path()
    # print(path)
    if os.path.exists(path):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Parts"
    header_fill = PatternFill("solid", start_color="2F4F8F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for col, h in enumerate(XLSX_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["H"].width = 30
    for letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[letter].width = 14
    wb.save(path)


def load_parts():
    ensure_xlsx()
    wb = load_workbook(get_xlsx_path())
    ws = wb.active
    parts = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        part_name = row[1] if len(row) > 1 else None
        if not part_name or str(part_name).strip() == "" or str(part_name).strip().lower() == "part":
            continue
        name = str(part_name).strip()
        parts[name] = {
            "elevation":   _to_num(row[3]),
            "cam_angle":   _to_num(row[4]),
            "table_angle": _to_num(row[5]),
            "zoom":        _to_num(row[6]),
            "focus":       _to_num(row[7]),
            "flash":       _to_num(row[8]),
        }
    return parts


def save_part(name, data):
    ensure_xlsx()
    path = get_xlsx_path()
    wb = load_workbook(path)
    ws = wb.active
    
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # Find existing row by Part name (col 2 = index 1)
    target_row = None
    for row in ws.iter_rows(min_row=2):
        if len(row) > 1 and str(row[1].value).strip() == name:
            target_row = row[1].row
            break

    if target_row is None:
        target_row = ws.max_row + 1
        # Auto-assign sr_no for new rows
        ws.cell(row=target_row, column=1, value=target_row - 1)

    # Write values into correct columns (1-based)
    col_values = [
        (2, data["elevation"]),
        (3, data["cam_angle"]),
        (4, data["table_angle"]),
        (5, data["zoom"]),
        (6, data["focus"]),
        (7, data["flash"]),
        (8, name),
    ]
    for col, val in col_values:
        cell = ws.cell(row=target_row, column=col, value=val)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    wb.save(path)


def _to_num(v):
    if v is None:
        return 0
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0
class MainWidget(QWidget):
    evtCallback = pyqtSignal(int)

    @staticmethod
    def makeLayout(lbl1, sli1, val1, lbl2, sli2, val2):
        hlyt1 = QHBoxLayout()
        hlyt1.addWidget(lbl1)
        hlyt1.addStretch()
        hlyt1.addWidget(val1)
        hlyt2 = QHBoxLayout()
        hlyt2.addWidget(lbl2)
        hlyt2.addStretch()
        hlyt2.addWidget(val2)
        vlyt = QVBoxLayout()
        vlyt.addLayout(hlyt1)
        vlyt.addWidget(sli1)
        vlyt.addLayout(hlyt2)
        vlyt.addWidget(sli2)
        return vlyt

    def __init__(self, output_folder=None):
        super().__init__()
        self.setWindowTitle("BFL Camera Control Panel")
        self.setMinimumSize(1200, 800)
        self.hcam = None
        self.imgWidth = 0
        self.imgHeight = 0
        self.pData = None
        self.frame = 0
        self.count = 0
        self.timer = QTimer(self)
        self.output_folder = output_folder
        self._loading_part = False  # guard to avoid re-entrant saves

        vlytctrl = QVBoxLayout()

        # ── Part selector ────────────────────────────────────────────────────
        part_group = QGroupBox("Part Settings")
        part_vlyt = QVBoxLayout()

        # Dropdown + management buttons
        hlyt_part = QHBoxLayout()
        self.combo_part = QComboBox()
        self.combo_part.setMinimumWidth(160)
        self.combo_part.currentIndexChanged.connect(self.onPartSelected)
        self.btn_add_part = QPushButton("+")
        self.btn_add_part.setFixedWidth(28)
        self.btn_add_part.setToolTip("Add new part")
        self.btn_add_part.clicked.connect(self.onAddPart)
        self.btn_save_part = QPushButton("Save")
        self.btn_save_part.setToolTip("Save current settings to this part")
        self.btn_save_part.clicked.connect(self.onSavePart)
        self.btn_reload_parts = QPushButton("↺")
        self.btn_reload_parts.setFixedWidth(28)
        self.btn_reload_parts.setToolTip("Reload parts from xlsx")
        self.btn_reload_parts.clicked.connect(self.refreshParts)
        hlyt_part.addWidget(QLabel("Part:"))
        hlyt_part.addWidget(self.combo_part, 1)
        hlyt_part.addWidget(self.btn_add_part)
        hlyt_part.addWidget(self.btn_save_part)
        hlyt_part.addWidget(self.btn_reload_parts)
        part_vlyt.addLayout(hlyt_part)

        # Three read-only info labels (Elevation, Cam Angle, Table Angle)
        info_hlyt = QHBoxLayout()
        self.lbl_elevation   = QLabel("Elevation: –")
        self.lbl_cam_angle   = QLabel("Cam Angle: –")
        self.lbl_table_angle = QLabel("Table Angle: –")
        for lbl in (self.lbl_elevation, self.lbl_cam_angle, self.lbl_table_angle):
            lbl.setAlignment(Qt.AlignCenter)
            info_hlyt.addWidget(lbl)
        part_vlyt.addLayout(info_hlyt)
        edit_hlyt = QHBoxLayout()
        self.edit_elevation   = QLineEdit("0")
        self.edit_cam_angle   = QLineEdit("0")
        self.edit_table_angle = QLineEdit("0")
        for field, placeholder in [
            (self.edit_elevation,   "Elevation"),
            (self.edit_cam_angle,   "Cam Angle"),
            (self.edit_table_angle, "Table Angle"),
        ]:
            field.setPlaceholderText(placeholder)
            field.setFixedWidth(80)
            field.textEdited.connect(self.onPartFieldEdited)
            edit_hlyt.addWidget(field)
        part_vlyt.addLayout(edit_hlyt)

        part_group.setLayout(part_vlyt)
        vlytctrl.addWidget(part_group)
        self.combo_camera = QComboBox()
        self.combo_camera.setMinimumWidth(300)
        self.btn_refresh = QPushButton("Refresh Cameras")
        self.btn_refresh.clicked.connect(self.refreshCameras)
        self.slider_flash = QSlider(Qt.Horizontal)
        self.slider_flash.setRange(0, 22)
        self.slider_flash.setValue(0)
        self.slider_flash.valueChanged.connect(self.onFlashChange)
        self.lbl_flash = QLabel("Flash: 0")
        self.edit_flash = QLineEdit("0")
        self.edit_flash.setFixedWidth(60)
        self.edit_flash.returnPressed.connect(self.onFlashEdit)
        hlyt_flash = QHBoxLayout()
        hlyt_flash.addWidget(self.lbl_flash)
        hlyt_flash.addWidget(self.slider_flash)
        hlyt_flash.addWidget(self.edit_flash)
        vlytctrl.addLayout(hlyt_flash)
        gboxexp = QGroupBox("Exposure")
        self.cbox_auto = QCheckBox("Auto exposure")
        self.cbox_auto.setEnabled(False)
        self.lbl_expoTime = QLabel("0")
        self.lbl_expoGain = QLabel("0")
        self.slider_expoTime = QSlider(Qt.Horizontal)
        self.slider_expoGain = QSlider(Qt.Horizontal)
        self.slider_expoTime.setEnabled(False)
        self.slider_expoGain.setEnabled(False)
        self.cbox_auto.stateChanged.connect(self.onAutoExpo)
        self.slider_expoTime.valueChanged.connect(self.onExpoTime)
        self.slider_expoGain.valueChanged.connect(self.onExpoGain)
        vlytexp = QVBoxLayout()
        vlytexp.addWidget(self.cbox_auto)
        vlytexp.addLayout(self.makeLayout(
            QLabel("Time:"), self.slider_expoTime, self.lbl_expoTime,
            QLabel("Gain:"), self.slider_expoGain, self.lbl_expoGain))
        gboxexp.setLayout(vlytexp)
        self.btn_autoWB = QPushButton("White balance")
        self.btn_autoWB.setEnabled(False)
        self.btn_autoWB.clicked.connect(self.onWB)
        self.btn_open = QPushButton("Open")
        self.btn_open.clicked.connect(self.onBtnOpen)
        self.btn_snap = QPushButton("Snap")
        self.btn_snap.setEnabled(False)
        self.btn_snap.clicked.connect(self.onBtnSnap)
        self.slider_zoom = QSlider(Qt.Horizontal)
        self.slider_zoom.setRange(5, 30)
        self.slider_zoom.setValue(15)
        self.slider_zoom.valueChanged.connect(self.onZoomChange)
        self.lbl_zoom = QLabel("Zoom: 1.5")
        self.edit_zoom = QLineEdit("1.5")
        self.edit_zoom.setFixedWidth(60)
        self.edit_zoom.returnPressed.connect(self.onZoomEdit)
        hlyt_zoom = QHBoxLayout()
        hlyt_zoom.addWidget(self.lbl_zoom)
        hlyt_zoom.addWidget(self.slider_zoom)
        hlyt_zoom.addWidget(self.edit_zoom)
        vlytctrl.addLayout(hlyt_zoom)

        self.btn_af_auto = QPushButton("Autofocus ON")
        self.btn_af_auto.setEnabled(False)
        self.btn_af_auto.clicked.connect(self.onAutoFocus)
        self.btn_af_off = QPushButton("Autofocus OFF")
        self.btn_af_off.setEnabled(False)
        self.btn_af_off.clicked.connect(self.onAutoFocusOff)
        vlytctrl.addWidget(self.btn_af_auto)
        vlytctrl.addWidget(self.btn_af_off)

        self.slider_focus = QSlider(Qt.Horizontal)
        self.slider_focus.setRange(0, 5068)
        self.slider_focus.setValue(0)
        self.slider_focus.valueChanged.connect(self.onFocusChange)
        self.lbl_focus = QLabel("Focus: 0")
        self.edit_focus = QLineEdit("0")
        self.edit_focus.setFixedWidth(60)
        self.edit_focus.returnPressed.connect(self.onFocusEdit)
        hlyt_focus = QHBoxLayout()
        hlyt_focus.addWidget(self.lbl_focus)
        hlyt_focus.addWidget(self.slider_focus)
        hlyt_focus.addWidget(self.edit_focus)
        vlytctrl.addLayout(hlyt_focus)

        vlytctrl.addWidget(self.combo_camera)
        vlytctrl.addWidget(self.btn_refresh)
        vlytctrl.addWidget(gboxexp)
        vlytctrl.addWidget(self.btn_autoWB)
        vlytctrl.addWidget(self.btn_open)
        vlytctrl.addWidget(self.btn_snap)
        vlytctrl.addStretch()
        wgctrl = QWidget()
        wgctrl.setLayout(vlytctrl)

        self.lbl_frame = QLabel()
        self.lbl_video = QLabel()
        vlytshow = QVBoxLayout()
        vlytshow.addWidget(self.lbl_video, 1)
        vlytshow.addWidget(self.lbl_frame)
        wgshow = QWidget()
        wgshow.setLayout(vlytshow)

        gmain = QGridLayout()
        gmain.setColumnStretch(0, 1)
        gmain.setColumnStretch(1, 4)
        gmain.addWidget(wgctrl)
        gmain.addWidget(wgshow)
        self.setLayout(gmain)

        self.evtCallback.connect(self.onevtCallback)
        self.timer.timeout.connect(self.onTimer)

        self.refreshCameras()
        self.refreshParts()

    def refreshParts(self):
        """Reload part list from xlsx and repopulate dropdown."""
        self._parts = load_parts()
        current = self.combo_part.currentText()
        self.combo_part.blockSignals(True)
        self.combo_part.clear()
        for name in self._parts:
            self.combo_part.addItem(name)
        # Restore previous selection if still present
        idx = self.combo_part.findText(current)
        self.combo_part.setCurrentIndex(max(idx, 0))
        self.combo_part.blockSignals(False)
        self.onPartSelected()

    def onPartSelected(self):
        """Populate edit fields and camera controls from the selected part."""
        name = self.combo_part.currentText()
        if not name or name not in self._parts:
            return
        d = self._parts[name]
        self._loading_part = True

        elev  = d["elevation"]
        cam   = d["cam_angle"]
        tbl   = d["table_angle"]
        zoom  = d["zoom"]
        focus = int(d["focus"])
        flash = int(d["flash"])

        self.lbl_elevation.setText(f"Elevation: {elev}")
        self.lbl_cam_angle.setText(f"Cam Angle: {cam}")
        self.lbl_table_angle.setText(f"Table Angle: {tbl}")

        self.edit_elevation.setText(str(elev))
        self.edit_cam_angle.setText(str(cam))
        self.edit_table_angle.setText(str(tbl))

        zoom_slider = max(5, min(30, int(float(zoom) * 10)))
        self.slider_zoom.setValue(zoom_slider)
        self.edit_zoom.setText(f"{float(zoom):.1f}")
        self.lbl_zoom.setText(f"Zoom: {float(zoom):.1f}")

        focus = max(0, min(5068, focus))
        self.slider_focus.setValue(focus)
        self.edit_focus.setText(str(focus))
        self.lbl_focus.setText(f"Focus: {focus}")

        flash = max(0, min(22, flash))
        self.slider_flash.setValue(flash)
        self.edit_flash.setText(str(flash))
        self.lbl_flash.setText(f"Flash: {flash}")

        if self.hcam is not None:
            try:
                self.hcam.put(uvcham.UVCHAM_ZOOM, zoom_slider)
            except Exception:
                pass
            try:
                self.hcam.put(uvcham.UVCHAM_AFPOSITION, focus)
            except Exception:
                pass
            try:
                self.hcam.put(uvcham.UVCHAM_LIGHT_ADJUSTMENT, flash)
            except Exception:
                pass

        self._loading_part = False

    def onPartFieldEdited(self):
        name = self.combo_part.currentText()
        if not name or self._loading_part:
            return
        elev  = self.edit_elevation.text()
        cam   = self.edit_cam_angle.text()
        tbl   = self.edit_table_angle.text()
        self.lbl_elevation.setText(f"Elevation: {elev}")
        self.lbl_cam_angle.setText(f"Cam Angle: {cam}")
        self.lbl_table_angle.setText(f"Table Angle: {tbl}")
        # Auto-save to xlsx
        self._save_current_part()

    def onSavePart(self):
        name = self.combo_part.currentText()
        if not name:
            QMessageBox.warning(self, "Warning", "No part selected.")
            return
        self._save_current_part()
        QMessageBox.information(self, "Saved",
            f"Settings for '{name}' saved to {XLSX_FILENAME}.")

    def _save_current_part(self):
        name = self.combo_part.currentText()
        if not name:
            return
        data = {
            "elevation":   _to_num(self.edit_elevation.text()),
            "cam_angle":   _to_num(self.edit_cam_angle.text()),
            "table_angle": _to_num(self.edit_table_angle.text()),
            "zoom":        _to_num(self.edit_zoom.text()),
            "focus":       _to_num(self.edit_focus.text()),
            "flash":       _to_num(self.edit_flash.text()),
        }
        self._parts[name] = data
        try:
            save_part(name, data)
        except Exception as e:
            print(f"Failed to save part: {e}")

    def onAddPart(self):
        """Prompt for a new part name and add it."""
        name, ok = QInputDialog.getText(self, "New Part", "Part name:")
        if not ok or not name.strip():
            return
        name = name.strip()
        if name in self._parts:
            QMessageBox.information(self, "Exists", f"'{name}' already exists.")
            self.combo_part.setCurrentText(name)
            return
        new_data = {"elevation": 0, "cam_angle": 0, "table_angle": 0,
                    "zoom": 1.5, "focus": 0, "flash": 0}
        self._parts[name] = new_data
        save_part(name, new_data)
        self.combo_part.addItem(name)
        self.combo_part.setCurrentText(name)

    def refreshCameras(self):
        self.combo_camera.clear()
        arr = uvcham.Uvcham.enum()
        for cam in arr:
            self.combo_camera.addItem(f"{cam.displayname} ({cam.id})", cam.id)

    def onBtnOpen(self):
        if self.hcam is not None:
            self.closeCamera()
        else:
            idx = self.combo_camera.currentIndex()
            if idx < 0:
                QMessageBox.warning(self, "Warning", "No camera selected.")
                return
            cam_id = self.combo_camera.itemData(idx)
            self.openCamera(cam_id)

    def onFlashChange(self, value):
        self.lbl_flash.setText(f"Flash: {value}")
        self.edit_flash.setText(str(value))
        if self.hcam is not None:
            self.hcam.put(uvcham.UVCHAM_LIGHT_ADJUSTMENT, value)
        if not self._loading_part:
            self._save_current_part()

    def onFlashEdit(self):
        val = self.edit_flash.text()
        try:
            ival = int(float(val))
            if 0 <= ival <= 22:
                self.slider_flash.setValue(ival)
                if self.hcam is not None:
                    self.hcam.put(uvcham.UVCHAM_LIGHT_ADJUSTMENT, ival)
        except Exception:
            pass

    def onBtnSnap(self):
        if self.hcam is not None and self.pData is not None:
            part_name = self.combo_part.currentText().strip()
            if not part_name:
                part_name = "unknown_part"
            image = QImage(self.pData, self.imgWidth, self.imgHeight, QImage.Format_RGB888)
            self.count += 1
            folder = getattr(self, 'output_folder', None)
            if folder is None:
                today = datetime.now().strftime('%Y-%m-%d')
                folder = os.path.join(os.getcwd(), today)
            if not os.path.exists(folder):
                os.makedirs(folder)
            fname = os.path.join(folder, f"{part_name}_{self.count}.jpg")
            image.save(fname)
            QMessageBox.information(self, "Saved", f"Image saved as {fname}")

    @staticmethod
    def eventCallBack(nEvent, self):
        self.evtCallback.emit(nEvent)

    def onevtCallback(self, nEvent):
        if self.hcam is not None:
            if uvcham.UVCHAM_EVENT_IMAGE & nEvent != 0:
                self.onImageEvent()
            elif uvcham.UVCHAM_EVENT_ERROR & nEvent != 0:
                self.closeCamera()
                QMessageBox.warning(self, "Warning", "Generic error.")
            elif uvcham.UVCHAM_EVENT_DISCONNECT & nEvent != 0:
                self.closeCamera()
                QMessageBox.warning(self, "Warning", "Camera disconnect.")

    def onImageEvent(self):
        self.hcam.pull(self.pData)
        self.frame += 1
        img = np.frombuffer(self.pData, dtype=np.uint8)
        try:
            if img.size == self.imgWidth * self.imgHeight * 3:
                image = QImage(self.pData, self.imgWidth, self.imgHeight, QImage.Format_RGB888)
            elif img.size == self.imgWidth * self.imgHeight * 3 // 2:
                img = img.reshape((int(self.imgHeight * 1.5), self.imgWidth))
                img_rgb = cv2.cvtColor(img, cv2.COLOR_YUV2RGB_I420)
                image = QImage(img_rgb.data, self.imgWidth, self.imgHeight, QImage.Format_RGB888)
            else:
                print(f'Buffer size mismatch: got {img.size}')
                return
            newimage = image.scaled(self.lbl_video.width(), self.lbl_video.height(),
                                    Qt.KeepAspectRatio, Qt.FastTransformation)
            self.lbl_video.setPixmap(QPixmap.fromImage(newimage))
        except Exception as e:
            print(f"Image decode error: {e}.")

    def onAutoExpo(self, state):
        if self.hcam is not None:
            self.hcam.put(uvcham.UVCHAM_AEXPO, 1 if state else 0)
            self.slider_expoTime.setEnabled(not state)
            self.slider_expoGain.setEnabled(not state)

    def onWB(self):
        if self.hcam is not None:
            self.hcam.put(uvcham.UVCHAM_WBMODE, 3)

    def onExpoTime(self, value):
        if self.hcam is not None:
            self.lbl_expoTime.setText(str(value))
            if not self.cbox_auto.isChecked():
                self.hcam.put(uvcham.UVCHAM_EXPOTIME, value)

    def onExpoGain(self, value):
        if self.hcam is not None:
            self.lbl_expoGain.setText(str(value))
            if not self.cbox_auto.isChecked():
                self.hcam.put(uvcham.UVCHAM_AGAIN, value)

    def updateExpoTime(self):
        val = self.hcam.get(uvcham.UVCHAM_EXPOTIME)
        with QSignalBlocker(self.slider_expoTime):
            self.slider_expoTime.setValue(val)
        self.lbl_expoTime.setText(str(val))

    def updateGain(self):
        val = self.hcam.get(uvcham.UVCHAM_AGAIN)
        with QSignalBlocker(self.slider_expoGain):
            self.slider_expoGain.setValue(val)

    def onTimer(self):
        if self.hcam is not None:
            self.lbl_frame.setText(str(self.frame))
            if self.cbox_auto.isChecked():
                self.updateExpoTime()
                self.updateGain()

    def onZoomChange(self, value):
        zoom_float = value / 10.0
        self.lbl_zoom.setText(f"Zoom: {zoom_float:.1f}")
        self.edit_zoom.setText(f"{zoom_float:.1f}")
        if self.hcam is not None:
            try:
                self.hcam.put(uvcham.UVCHAM_ZOOM, int(zoom_float * 10))
            except Exception as e:
                print(f'Failed to set zoom: {e}')
        if not self._loading_part:
            self._save_current_part()

    def onZoomEdit(self):
        val = self.edit_zoom.text()
        try:
            fval = float(val)
            if 0.5 <= fval <= 3.0:
                self.slider_zoom.setValue(int(fval * 10))
                if self.hcam is not None:
                    self.hcam.put(uvcham.UVCHAM_ZOOM, int(fval * 10))
        except Exception:
            pass

    def onFocusChange(self, value):
        self.lbl_focus.setText(f"Focus: {value}")
        self.edit_focus.setText(str(value))
        if self.hcam is not None:
            try:
                self.hcam.put(uvcham.UVCHAM_AFPOSITION, value)
            except Exception as e:
                print(f'Failed to set manual focus: {e}')
        if not self._loading_part:
            self._save_current_part()

    def onFocusEdit(self):
        val = self.edit_focus.text()
        try:
            ival = int(float(val))
            if 0 <= ival <= 5068:
                self.slider_focus.setValue(ival)
                if self.hcam is not None:
                    self.hcam.put(uvcham.UVCHAM_AFPOSITION, ival)
        except Exception:
            pass

    def onAutoFocus(self):
        if self.hcam is not None:
            try:
                self.hcam.put(uvcham.UVCHAM_AFMODE, 1)
                self.btn_af_auto.setEnabled(False)
                self.btn_af_off.setEnabled(True)
            except Exception as e:
                print(f'Failed to enable autofocus: {e}')

    def onAutoFocusOff(self):
        if self.hcam is not None:
            try:
                self.hcam.put(uvcham.UVCHAM_AFMODE, 0)
                self.btn_af_auto.setEnabled(True)
                self.btn_af_off.setEnabled(False)
            except Exception as e:
                print(f'Failed to disable autofocus: {e}')

    def openCamera(self, id):
        self.hcam = uvcham.Uvcham.open(id)
        if self.hcam:
            self.frame = 0
            self.hcam.put(uvcham.UVCHAM_FORMAT, 2)
            res = self.hcam.get(uvcham.UVCHAM_RES)
            self.imgWidth = self.hcam.get(uvcham.UVCHAM_WIDTH | res)
            self.imgHeight = self.hcam.get(uvcham.UVCHAM_HEIGHT | res)
            self.pData = bytes(uvcham.TDIBWIDTHBYTES(self.imgWidth * 24) * self.imgHeight)
            try:
                self.hcam.start(None, self.eventCallBack, self)
            except uvcham.HRESULTException:
                self.closeCamera()
                QMessageBox.warning(self, "Warning", "Failed to start camera.")
            else:
                self.cbox_auto.setEnabled(True)
                self.btn_autoWB.setEnabled(True)
                self.btn_open.setText("Close")
                self.btn_snap.setEnabled(True)
                self.slider_flash.setEnabled(True)
                self.btn_af_auto.setEnabled(True)
                self.btn_af_off.setEnabled(True)
                self.slider_zoom.setEnabled(True)
                self.slider_focus.setEnabled(True)
                nmin, nmax, ndef = self.hcam.range(uvcham.UVCHAM_EXPOTIME)
                self.slider_expoTime.setRange(nmin, nmax)
                nmin, nmax, ndef = self.hcam.range(uvcham.UVCHAM_AGAIN)
                self.slider_expoGain.setRange(nmin, nmax)
                bAuto = self.hcam.get(uvcham.UVCHAM_AEXPO)
                self.cbox_auto.setChecked(1 == bAuto)
                self.slider_expoTime.setEnabled(1 != bAuto)
                self.slider_expoGain.setEnabled(1 != bAuto)
                self.updateExpoTime()
                self.updateGain()
                self.timer.start(1000)
                # Apply current part settings to newly opened camera
                self.onPartSelected()

    def closeCamera(self):
        if self.hcam:
            self.hcam.put(uvcham.UVCHAM_LIGHT_ADJUSTMENT, 0)
            self.hcam.put(uvcham.UVCHAM_AFMODE, 0)
            self.hcam.close()
        self.hcam = None
        self.pData = None
        self.btn_open.setText("Open")
        self.timer.stop()
        self.lbl_frame.clear()
        self.cbox_auto.setEnabled(False)
        self.slider_expoGain.setEnabled(False)
        self.slider_expoTime.setEnabled(False)
        self.btn_autoWB.setEnabled(False)
        self.btn_snap.setEnabled(False)
        self.slider_flash.setEnabled(False)
        self.btn_af_auto.setEnabled(False)
        self.btn_af_off.setEnabled(False)
        self.slider_zoom.setEnabled(False)
        self.slider_focus.setEnabled(False)

    def closeEvent(self, event):
        if self.hcam is not None:
            self.hcam.close()
            self.hcam = None


if __name__ == '__main__':
    app = QApplication(sys.argv)
    output_folder = None
    if len(sys.argv) > 1:
        output_folder = sys.argv[1]
    mw = MainWidget(output_folder=output_folder)
    mw.show()
    sys.exit(app.exec_())